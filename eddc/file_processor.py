"""
Moduł do obsługi plików Excel z użyciem openpyxl, z obsługą błędów API i kolorowaniem wierszy.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from eddc.api import process_driver_data
import time


def update_xlsx(input_file_path, output_file_path):
    """
    Aktualizuje plik Excel danymi z API i koloruje wiersze w zależności od statusu.

    :param input_file_path: Ścieżka do pliku Excel.
    """
    wb = load_workbook(input_file_path)
    sheet = wb.active

    # Nagłówki i ich indeksy (zakładamy, że wiersz 1 to nagłówki)
    headers = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}

    # Dodanie brakujących kolumn, jeśli ich brakuje
    required_columns = ["wydający", "ważność", "status", "zmiana statusu"]
    for column in required_columns:
        if column not in headers:
            headers[column] = max(value for key, value in headers.items() if key is not None) + 1
            sheet.cell(row=1, column=headers[column]).value = column

    # Definicje kolorów
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")    # Light red

    # Przetwarzanie wierszy
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False), start=2):
        first_name = row[headers["imię"] - 1].value
        last_name = row[headers["nazwisko"] - 1].value
        document_number = row[headers["nr dokumentu"] - 1].value

        result = {}

        # Wywołanie API
        result = process_driver_data(first_name, last_name, document_number)
        time.sleep(0.1)
        if "dokumentPotwierdzajacyUprawnienia" in result:
            doc = result["dokumentPotwierdzajacyUprawnienia"]
            sheet.cell(row=row_index, column=headers["wydający"]).value = doc.get("organWydajacyDokument", {}).get("wartosc", "brak danych")
            sheet.cell(row=row_index, column=headers["ważność"]).value = doc.get("dataWaznosci", "Bezterminowo")
            sheet.cell(row=row_index, column=headers["status"]).value = doc.get("stanDokumentu", {}).get("stanDokumentu", {}).get("wartosc", "brak danych")
            sheet.cell(row=row_index, column=headers["zmiana statusu"]).value = "; ".join(
                doc.get("stanDokumentu", {}).get("powodZmianyStanu", [])
            )
            fill = green_fill
        else:
            # Obsługa błędów
            error_message = result.get("error", "brak danych")
            sheet.cell(row=row_index, column=headers["status"]).value = error_message
            fill = red_fill

        # Kolorowanie wiersza
        for cell in row:
            cell.fill = fill

        print(f"{row_index-1}. {document_number.upper()}: {'OK' if fill == green_fill else 'NIE OK'}")

    # Zapisanie pliku
    if output_file_path:
        wb.save(output_file_path)
