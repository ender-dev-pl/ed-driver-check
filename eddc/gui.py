"""
Moduł GUI do obsługi interaktywnej aplikacji.
"""

import tkinter as tk
from tkinter import filedialog
from tkinter import ttk, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from eddc.api import process_driver_data


def run_gui():
    """
    Uruchamia interfejs graficzny.
    """
    loaded_file = {"path": None}  # Przechowywanie załadowanego pliku

    def load_file():
        """
        Załaduj plik Excel i wyświetl wstępne dane w tabeli.
        """
        file_path = filedialog.askopenfilename(filetypes=[("Pliki Excel", "*.xlsx")])
        if not file_path:
            return

        try:
            # Załaduj dane do tabeli
            clear_table()
            wb = load_workbook(file_path)
            sheet = wb.active
            loaded_file["path"] = file_path

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                first_name = row[0] or ""
                last_name = row[1] or ""
                document_number = row[2] or ""

                driver_name = f"{first_name.capitalize()} {last_name[0].capitalize()}." if first_name and last_name else "Brak danych"
                document_number = document_number.upper()

                add_row(row_index, driver_name, document_number, 6 * [], color=None)

            messagebox.showinfo("Sukces", f"Załadowano plik: {file_path}")
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się załadować pliku: {str(e)}")

    def process_file():
        """
        Przetwórz załadowany plik Excel, wywołaj API dla każdego wiersza i zapisz wyniki.
        """
        if not loaded_file["path"]:
            messagebox.showwarning("Brak pliku", "Najpierw załaduj plik Excel.")
            return

        try:
            wb = load_workbook(loaded_file["path"])
            sheet = wb.active
            #clear_table()

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=False), start=2):
                first_name = row[0].value or ""
                last_name = row[1].value or ""
                document_number = row[2].value or ""

                result = process_driver_data(first_name, last_name, document_number)
                root.after(100)

                driver_name = f"{first_name.capitalize()} {last_name[0].capitalize()}." if first_name and last_name else "Brak danych"
                document_number = document_number.upper()

                if "dokumentPotwierdzajacyUprawnienia" in result:
                    doc = result["dokumentPotwierdzajacyUprawnienia"]
                    status = doc.get("stanDokumentu", {}).get("stanDokumentu", {}).get("wartosc", "")
                    color = "90EE90" if status == "Wydany" else "FFCCCB"
                else:
                    doc = {}
                    status = "Błąd danych"
                    color = "FFCCCB"

                # zaktualizuj wiersz w tabeli
                update_row(row_index - 2, driver_name, document_number,
                        doc.get("typDokumentu", {}).get("wartosc", "brak danych"),
                        doc.get("seriaNumerBlankietuDruku", "brak danych"),
                        doc.get("organWydajacyDokument", {}).get("wartosc", "brak danych"),
                        doc.get("dataWaznosci", "Bezterminowo" if doc else ""),
                        status,
                        "; ".join([f"{k['kategoria']} ({k['dataWaznosci'] or 'Bezterminowo'})"
                                   for k in doc.get("daneUprawnieniaKategorii", [])]),
                        color="lightgreen" if status == "Wydany" else "lightcoral")

                root.update()

        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się przetworzyć pliku: {str(e)}")



    def check_driver_permissions():
        """
        Sprawdź uprawnienia kierowcy na podstawie danych w formularzu.
        """
        first_name = entry_first_name.get()
        last_name = entry_last_name.get()
        document_number = entry_doc_number.get()

        result = process_driver_data(first_name, last_name, document_number)

        if "dokumentPotwierdzajacyUprawnienia" in result:
            doc = result["dokumentPotwierdzajacyUprawnienia"]
            driver_name = f"{first_name.capitalize()} {last_name[0].capitalize()}." if first_name and last_name else "Brak danych"
            document_number = document_number.upper()

            color = "lightgreen" if doc.get("stanDokumentu", {}).get("stanDokumentu", {}).get("wartosc") == "Wydany" else "lightcoral"
            add_row("", driver_name, document_number,
                    doc.get("typDokumentu", {}).get("wartosc", "brak danych"),
                    doc.get("seriaNumerBlankietuDruku", "brak danych"),
                    doc.get("organWydajacyDokument", {}).get("wartosc", "brak danych"),
                    doc.get("dataWaznosci", "Bezterminowo"),
                    doc.get("stanDokumentu", {}).get("stanDokumentu", {}).get("wartosc", "brak danych"),
                    "; ".join([f"{k['kategoria']} ({k['dataWaznosci'] or 'Bezterminowo'})"
                               for k in doc.get("daneUprawnieniaKategorii", [])]),
                    color=color)
        else:
            messagebox.showinfo("Błąd", result.get("error", "brak danych"))

    def clear_table():
        """
        Wyczyść tabelę wyników.
        """
        for row in tree.get_children():
            tree.delete(row)

    def add_row(index, driver_name, document_number, *values, color=None):
        """
        Dodaj wiersz do tabeli z opcjonalnym kolorowaniem.

        :param index: Numer wiersza.
        :param driver_name: Pole "Kierowca".
        :param document_number: Numer dokumentu (wielkie litery).
        :param values: Pozostałe kolumny.
        :param color: Kolor tła wiersza.
        """
        row_id = tree.insert("", "end", values=(index, driver_name, document_number, *values))
        if color:
            tree.tag_configure(color, background=color)
            tree.item(row_id, tags=(color,))

    def update_row(index, driver_name, document_number, *values, color=None):
        """
        Zaktualizuj wiersz w tabeli.

        :param index: Numer wiersza w tabeli (zaczyna się od 0).
        :param driver_name: Pole "Kierowca".
        :param document_number: Numer dokumentu (wielkie litery).
        :param values: Pozostałe kolumny.
        :param color: Kolor tła wiersza.
        """
        # Pobierz ID wiersza na podstawie indeksu
        row_id = tree.get_children()[index]

        # Zaktualizuj wartości w wierszu
        tree.item(row_id, values=(index + 1, driver_name, document_number, *values))

        # Zaktualizuj kolor tła wiersza
        if color:
            tree.tag_configure(color, background=color)
            tree.item(row_id, tags=(color,))


    # Konfiguracja okna głównego
    root = tk.Tk()
    root.title("ED-Driver-Check - Sprawdzanie uprawnień kierowców by EnderDEV")

    # Konfiguracja stylu dla ramek
    style = ttk.Style()
    style.configure("TFrame", relief="groove", padding=10)  # Ustawienie stylu dla TFrame

    # Kontener na formularz i przycisk
    form_frame = ttk.LabelFrame(root, style="TFrame", text="Sprawdzenie pojedynczego kierowcy")
    form_frame.grid(row=0, column=0, columnspan=2, pady=5, padx=5 )

    # Formularz w kontenerze
    ttk.Label(form_frame, text="Imię:").grid(row=0, column=0, sticky="e")
    entry_first_name = ttk.Entry(form_frame)
    entry_first_name.grid(row=0, column=1, padx=5, pady=5, sticky="we")

    ttk.Label(form_frame, text="Nazwisko:").grid(row=1, column=0, sticky="ee")
    entry_last_name = ttk.Entry(form_frame)
    entry_last_name.grid(row=1, column=1, padx=5, pady=5, sticky="we")

    ttk.Label(form_frame, text="Numer dokumentu:").grid(row=2, column=0, padx=(10,0), sticky="e")
    entry_doc_number = ttk.Entry(form_frame)
    entry_doc_number.grid(row=2, column=1, padx=5, pady=5, sticky="we")

    # Przycisk po prawej stronie formularza
    ttk.Button(form_frame, text="Sprawdź uprawnienia", command=check_driver_permissions).grid(row=1, column=2, rowspan=1, padx=10, pady=5, sticky="ns")

    # Kontener na przyciski
    button_frame = ttk.LabelFrame(root, style="TFrame", text = "Sprawdzenie zbiorowe")
    button_frame.grid(row=6, column=0, columnspan=2, pady=5, padx=5)

    # Przycisk do załadowania pliku Excel
    ttk.Button(button_frame, text="Załaduj plik Excel", command=load_file).pack(side="left", padx=5)

    # Przycisk do przetworzenia pliku
    ttk.Button(button_frame, text="Przetwórz plik", command=process_file).pack(side="left", padx=5)


    # Tabela wyników
    columns = ("Nr", "Kierowca", "Nr dokumentu", "Typ dokumentu", "Seria i numer", "Organ wydający", "Ważność", "Status dokumentu", "Kategorie")
    tree = ttk.Treeview(root, columns=columns, show="headings", height=10)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=150, anchor="center")

    tree.grid(row=7, column=0, columnspan=2, sticky="nsew", pady=10)

    scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=7, column=2, sticky="ns")

    # Ustawienia okna
    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(7, weight=1)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
